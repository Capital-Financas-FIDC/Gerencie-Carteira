#Mantém todas as funcionalidades da versão anterior

#Bug fixes:
#   -Conserta o problema de que caso não houvesse uma planilha na pasta Diário com a data exatamente do dia anterior
#    o programa falhava. Isso costumava acontecer nos finais de semana, quando ficávamos um dia ou dois sem receber 
#    um relatório do Gerencie Carteira. Agora o programa detecta a planilha com a data mais recente, ao invés de uma
#    com a data exatamente anterior.


import win32com.client
import os
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re

# Conectar ao Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# Acessar a caixa de entrada
inbox = outlook.GetDefaultFolder(6)  # 6 corresponde à pasta Inbox
messages = inbox.Items

# Definir o assunto do e-mail para pesquisa
assunto_procurado = "Gerencie Carteira - Consulte as Empresas Monitoradas"

# Ordenar e-mails por data, do mais antigo ao mais recente
messages.Sort("[ReceivedTime]", False)

# Filtrar apenas os e-mails não lidos com o assunto especificado
emails_nao_lidos = [msg for msg in messages if msg.Subject == assunto_procurado and msg.UnRead]

if emails_nao_lidos:
    pasta_destino = r"C:\Users\comercial05\Documents\Gerencie Carteira\HTML"
    dados = []

    # Extrair a data do e-mail mais recente
    data_mais_recente = max(email.ReceivedTime for email in emails_nao_lidos)

    # Calcular a data da última planilha
    pasta_diario = r"C:\Users\comercial05\Documents\Gerencie Carteira\Diário" #Definir a pasta das planilhas diárias

    arquivos_diario = os.listdir(pasta_diario) #Lista todos os arquivos na pasta

    padrao_nome_excel = re.compile(r"Gerencie Carteira_(\d{2}_\d{2}_\d{4})") #Filtrar os arquivos na pasta que têm o padrão de nome edsejado
    
    datas_encontradas_lista = []

    for arquivo_diario in arquivos_diario:
        match = padrao_nome_excel.search(arquivo_diario)
        if match:
            datas_encontradas_lista.append(match.group(1))
        else:
            print("Nenhum arquivo com data foi encontrado na pasta!")
            break

        #Se encontrar alguma data, pegar a mais recente
        if datas_encontradas_lista:
            ultima_data_diario = max(datas_encontradas_lista, key=lambda x: tuple(map(int, x.split('_')))[::-1])
            #print(ultima_data_diario) # Apenas para depuração
            nome_arquivo_excel = f"Gerencie Carteira_{ultima_data_diario}.xlsx" #Declara o nome correto da útlima planilha na pasta
            caminho_excel = os.path.join(pasta_diario, nome_arquivo_excel) #Define qual é o endereço correto do último excel que está na pasta
        else:
            print("⚠ Nenhuma planilha encontrada na pasta! Verifique manualmente")

    #print(f" Usando o arquivo: {caminho_excel}: ")  # Apenas para depuração

    for email in emails_nao_lidos:
        # Extrair a data do e-mail e formatá-la
        data_email = email.ReceivedTime.strftime("%Y/%m/%d")

        # Baixar o anexo .html
        caminho_arquivo = None
        for anexo in email.Attachments:
            if anexo.FileName.endswith(".html"):
                nome_arquivo = f"Gerencie_Carteira_{data_email.replace('/', '_')}.html"
                caminho_arquivo = os.path.join(pasta_destino, nome_arquivo)
                anexo.SaveAsFile(caminho_arquivo)
                print(f"✅ Anexo salvo como: '{nome_arquivo}' na pasta HTML's do Gerencie Carteira")
                break

        if caminho_arquivo:
            # Abrir o arquivo HTML e extrair os dados da tabela
            with open(caminho_arquivo, "r", encoding="utf-8") as arquivo:
                soup = BeautifulSoup(arquivo, "html.parser")

            tabelas = soup.find_all("table")
            if len(tabelas) > 1:
                tabela = tabelas[1]
                tbody = tabela.find("tbody")
                linhas = tbody.find_all("tr") if tbody else tabela.find_all("tr")

                for linha in linhas:
                    colunas = linha.find_all("td")
                    if len(colunas) >= 3:
                        cnpj = colunas[0].get_text(strip=True)
                        razao_social = colunas[1].get_text(strip=True)
                        alteracao = colunas[2].get_text(strip=True)

                        if "CNPJ" not in cnpj and "Razão Social" not in razao_social and "Alteração" not in alteracao:
                            dados.append([cnpj, razao_social, alteracao, data_email])

    if dados:
        # Criar um DataFrame do pandas com a nova coluna "Data da Operação"
        df = pd.DataFrame(dados, columns=["CNPJ", "Razão Social", "Alteração", "Data da Operação"])

        # Garantir que os valores sejam strings
        df["CNPJ"] = df["CNPJ"].astype(str)
        df["Razão Social"] = df["Razão Social"].astype(str)
        df["Alteração"] = df["Alteração"].astype(str)

        # Adicionar um espaço não-quebrante e um espaço convencional
        df["CNPJ"] = df["CNPJ"].apply(lambda x: "\xa0 " + x if isinstance(x, str) else x)
        df["Razão Social"] = df["Razão Social"].apply(lambda x: "\xa0 " + x if isinstance(x, str) else x)
        df["Alteração"] = df["Alteração"].apply(lambda x: "\xa0 " + x if isinstance(x, str) else x)

        # Ordenar os dados do mais antigo para o mais novo
        df = df.sort_values(by="Data da Operação", ascending=True)

        # Remover colunas desnecessárias e garantir que todas as colunas sejam mantidas
        df = df.loc[:, ["CNPJ", "Razão Social", "Alteração", "Data da Operação"]]

        print(df)

        # Carregar o arquivo Excel e acessar a segunda planilha diretamente
        wb = load_workbook(caminho_excel)
        ws_email_bd = wb.worksheets[1]  # A segunda aba, garantindo que seja "E-Mail BD"
        ws_procv_gerentes_bd = wb.worksheets[3]  # A quarta aba, garantindo que seja "PROCV GERENTES BD"

        # Verificar se o nome da planilha corresponde ao esperado
        if ws_email_bd.title != "E-Mail BD":
            print("⚠ A segunda planilha não tem o nome esperado. Verifique manualmente.")
        elif ws_procv_gerentes_bd.title != "PROCV GERENTES BD":
            print("⚠ A quarta planilha não tem o nome esperado. Verifique manualmente.")
        else:
            # Encontrar a primeira linha vazia na planilha "E-Mail BD"
            primeira_linha_vazia_email_bd = ws_email_bd.max_row + 1

            # Inserir os novos dados na planilha "E-Mail BD"
            for r_idx, row in enumerate(df.values, start=primeira_linha_vazia_email_bd):
                for c_idx, value in enumerate(row, start=1):
                    ws_email_bd.cell(row=r_idx, column=c_idx, value=value)

            # **Preencher automaticamente as células E, F, G e H COM REFERÊNCIA DINÂMICA CORRETA**
            for coluna in ["E", "F", "G", "H"]:  # Percorre cada coluna
                ultima_linha_preenchida_email_bd = primeira_linha_vazia_email_bd - 1  # Última linha antes de colar os novos dados
                formula_origem_email_bd = ws_email_bd[f"{coluna}{ultima_linha_preenchida_email_bd}"].value  # Captura a fórmula original da célula acima

                # Verifica se a célula contém uma fórmula
                if formula_origem_email_bd and isinstance(formula_origem_email_bd, str) and formula_origem_email_bd.startswith("="):  
                    for linha in range(primeira_linha_vazia_email_bd, ws_email_bd.max_row + 1):
                        nova_formula = re.sub(r'([A-Z])(\d+)', lambda match: f"{match.group(1)}{linha}" if match.group(2) == str(ultima_linha_preenchida_email_bd) else match.group(0), formula_origem_email_bd)
                        ws_email_bd[f"{coluna}{linha}"].value = nova_formula  # Aplica a fórmula ajustada para a linha correta

                else:
                    for linha in range(primeira_linha_vazia_email_bd, ws_email_bd.max_row + 1):
                        ws_email_bd[f"{coluna}{linha}"].value = f"={coluna}{linha-1}"  

            # Aplicar alinhamento à coluna D (Data da Operação)
            for cell in ws_email_bd["D"]:
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # Verificar se cada informação adicionada na coluna B está presente na coluna A da planilha "PROCV GERENTES BD"
            valores_procv_gerentes_bd = [cell.value for cell in ws_procv_gerentes_bd['A']]
            erros_detectados = False #Flag para encontrar erros

            for linha in range(primeira_linha_vazia_email_bd, ws_email_bd.max_row + 1):
                valor_coluna_b = ws_email_bd[f"B{linha}"].value
                if valor_coluna_b not in valores_procv_gerentes_bd:
                    print(f"⚠ Erro encontrado na célula E{linha}. Verifique manualmente.")
                    erros_detectados = True #Caso um erro seja detectado, define a flag como True

            # Se algum erro for detectado, prossegue normalmente e salva as alterações no arquivo Excel
            if erros_detectados==True:
                novo_nome_arquivo = f"Gerencie Carteira_{data_email.replace('/', '_')}.xlsx"
                caminho_novo_excel = os.path.join(r"C:\Users\comercial05\Documents\Gerencie Carteira\Diário", novo_nome_arquivo)
                wb.save(caminho_novo_excel)
                print(f"✅ Arquivo salvo como: '{novo_nome_arquivo}' na pasta Diário do Gerencie Carteira")

            #Se nenhum erro for detectado, atualiza a tabela dinâmica
            elif erros_detectados==False:
                try:
                    ws_tabela_dinamica = wb.worksheets[0] # A primeira aba, garantindo que seja "Tabela Dinâmica"
                    #Atualiza a tabela dinâmica
                    pivot = ws_tabela_dinamica._pivots[0]
                    pivot.cache.refreshOnLoad = True
                    novo_nome_arquivo = f"Gerencie Carteira_{data_email.replace('/', '_')}.xlsx"
                    caminho_novo_excel = os.path.join(r"C:\Users\comercial05\Documents\Gerencie Carteira\Diário", novo_nome_arquivo)
                    wb.save(caminho_novo_excel)
                    print(f"✅ Tabela dinâmica atualizada com sucesso!")
                    print(f"✅ Arquivo salvo como: '{novo_nome_arquivo}' na pasta Diário do Gerencie Carteira")
                except Exception as e:
                    print(f"❌ Erro ao tentar atualizar a tabela dinâmica: {e}")
    else:
        print("⚠ Nenhuma tabela válida encontrada nos arquivos HTML.")
else:
    print("⚠ Nenhum e-mail não lido encontrado com o assunto especificado.")