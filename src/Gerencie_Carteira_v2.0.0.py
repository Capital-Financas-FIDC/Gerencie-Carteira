#Este programa identifica o email correto, baixa o html anexado na pasta correta,
#copia os dados e cola no excel da base de dados do gerencie carteira corretamente


import win32com.client
import os
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Caminho correto do Excel
caminho_excel = r"C:\Users\comercial05\Documents\Gerencie Carteira\Diário\Gerencie Carteira 25.05.07.xlsx"

# Conectar ao Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# Acessar a caixa de entrada
inbox = outlook.GetDefaultFolder(6)  # 6 corresponde à pasta Inbox
messages = inbox.Items

# Definir o assunto do e-mail para pesquisa
assunto_procurado = "Gerencie Carteira - Consulte as Empresas Monitoradas"

# Ordenar e-mails por data, do mais antigo ao mais recente
messages.Sort("[ReceivedTime]", False)

# Filtrar apenas os e-mails não lidos com o assunto especificado
emails_nao_lidos = [msg for msg in messages if msg.Subject == assunto_procurado and msg.UnRead]

if emails_nao_lidos:
    pasta_destino = r"C:\Users\comercial05\Documents\Gerencie Carteira\HTML"
    dados = []

    for email in emails_nao_lidos:
        # Extrair a data do e-mail e formatá-la
        data_email = email.ReceivedTime.strftime("%Y/%m/%d")

        # Baixar o anexo .html
        caminho_arquivo = None
        for anexo in email.Attachments:
            if anexo.FileName.endswith(".html"):
                nome_arquivo = f"Gerencie_Carteira_{data_email.replace('/', '_')}.html"
                caminho_arquivo = os.path.join(pasta_destino, nome_arquivo)
                anexo.SaveAsFile(caminho_arquivo)
                print(f"Anexo salvo em: {caminho_arquivo}")
                break

        if caminho_arquivo:
            # Abrir o arquivo HTML e extrair os dados da tabela
            with open(caminho_arquivo, "r", encoding="utf-8") as arquivo:
                soup = BeautifulSoup(arquivo, "html.parser")

            tabelas = soup.find_all("table")
            if len(tabelas) > 1:
                tabela = tabelas[1]
                tbody = tabela.find("tbody")
                linhas = tbody.find_all("tr") if tbody else tabela.find_all("tr")

                for linha in linhas:
                    colunas = linha.find_all("td")
                    if len(colunas) >= 3:
                        cnpj = colunas[0].get_text(strip=True)
                        razao_social = colunas[1].get_text(strip=True)
                        alteracao = colunas[2].get_text(strip=True)

                        if "CNPJ" not in cnpj and "Razão Social" not in razao_social and "Alteração" not in alteracao:
                            dados.append([cnpj, razao_social, alteracao, data_email])

    if dados:
        # Criar um DataFrame do pandas com a nova coluna "Data da Operação"
        df = pd.DataFrame(dados, columns=["CNPJ", "Razão Social", "Alteração", "Data da Operação"])

        # Garantir que os valores sejam strings
        df["CNPJ"] = df["CNPJ"].astype(str)
        df["Razão Social"] = df["Razão Social"].astype(str)
        df["Alteração"] = df["Alteração"].astype(str)

        # Adicionar um espaço não-quebrante e um espaço convencional
        df["CNPJ"] = df["CNPJ"].apply(lambda x: "\xa0 " + x if isinstance(x, str) else x)
        df["Razão Social"] = df["Razão Social"].apply(lambda x: "\xa0 " + x if isinstance(x, str) else x)
        df["Alteração"] = df["Alteração"].apply(lambda x: "\xa0 " + x if isinstance(x, str) else x)

        # Ordenar os dados do mais antigo para o mais novo
        df = df.sort_values(by="Data da Operação", ascending=True)

        # Remover colunas desnecessárias e garantir que todas as colunas sejam mantidas
        df = df.loc[:, ["CNPJ", "Razão Social", "Alteração", "Data da Operação"]]

        print(df)

        # Carregar o arquivo Excel e acessar a segunda planilha diretamente
        wb = load_workbook(caminho_excel)

        # Selecionar a segunda planilha baseada na posição, garantindo que seja "E-Mail BD"
        ws = wb.worksheets[1]  # Segunda aba (índice começa do 0)

        # Verificar se o nome da planilha corresponde ao esperado
        if ws.title != "E-Mail BD":
            print("⚠ A segunda planilha não tem o nome esperado. Verifique manualmente.")
        else:
            # Encontrar a primeira linha vazia
            primeira_linha_vazia = ws.max_row + 1

            # Inserir os dados na primeira linha disponível
            for r_idx, row in enumerate(df.values, start=primeira_linha_vazia):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Aplicar alinhamento à coluna D (Data da Operação)
            for cell in ws["D"]:
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # Salvar as alterações no arquivo Excel
            wb.save(caminho_excel)
            print(f"✅ Dados copiados corretamente para a planilha 'E-Mail BD' ({ws.title}) na primeira linha disponível!")
    else:
        print("Nenhuma tabela válida encontrada nos arquivos HTML.")
else:
    print("Nenhum e-mail não lido encontrado com o assunto especificado.")