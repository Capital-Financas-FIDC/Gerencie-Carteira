#Esta versão busca e encontra o email corretamente, baixa e salva o html da pasta correta,
#copia e cola os dados corretamente no banco de dados do gerencie carteira, puxa com referência dinâmica corretamente,
#e salva num novo arquivo nomeado corretamente

import win32com.client
import os
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re

# Conectar ao Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# Acessar a caixa de entrada
inbox = outlook.GetDefaultFolder(6)  # 6 corresponde à pasta Inbox
messages = inbox.Items

# Definir o assunto do e-mail para pesquisa
assunto_procurado = "Gerencie Carteira - Consulte as Empresas Monitoradas"

# Ordenar e-mails por data, do mais antigo ao mais recente
messages.Sort("[ReceivedTime]", False)

# Filtrar apenas os e-mails não lidos com o assunto especificado
emails_nao_lidos = [msg for msg in messages if msg.Subject == assunto_procurado and msg.UnRead]

if emails_nao_lidos:
    pasta_destino = r"C:\Users\comercial05\Documents\Gerencie Carteira\HTML"
    dados = []

    # Extrair a data do e-mail mais recente
    data_mais_recente = max(email.ReceivedTime for email in emails_nao_lidos)

    # Calcular a data do dia anterior
    data_anterior = data_mais_recente - timedelta(days=1)

    # Formatar a data no padrão correto para o nome do Excel (yyyy.mm.dd)
    nome_arquivo_excel = f"Gerencie Carteira_{data_anterior.strftime('%Y_%m_%d')}.xlsx"

    # Definir o caminho correto do Excel dinamicamente
    caminho_excel = os.path.join(r"C:\Users\comercial05\Documents\Gerencie Carteira\Diário", nome_arquivo_excel)

    print(f" Usando o arquivo: {caminho_excel}: ")  # Apenas para depuração

    for email in emails_nao_lidos:
        # Extrair a data do e-mail e formatá-la
        data_email = email.ReceivedTime.strftime("%Y/%m/%d")

        # Baixar o anexo .html
        caminho_arquivo = None
        for anexo in email.Attachments:
            if anexo.FileName.endswith(".html"):
                nome_arquivo = f"Gerencie_Carteira_{data_email.replace('/', '_')}.html"
                caminho_arquivo = os.path.join(pasta_destino, nome_arquivo)
                anexo.SaveAsFile(caminho_arquivo)
                print(f"✅ Anexo salvo em: {caminho_arquivo}")
                break

        if caminho_arquivo:
            # Abrir o arquivo HTML e extrair os dados da tabela
            with open(caminho_arquivo, "r", encoding="utf-8") as arquivo:
                soup = BeautifulSoup(arquivo, "html.parser")

            tabelas = soup.find_all("table")
            if len(tabelas) > 1:
                tabela = tabelas[1]
                tbody = tabela.find("tbody")
                linhas = tbody.find_all("tr") if tbody else tabela.find_all("tr")

                for linha in linhas:
                    colunas = linha.find_all("td")
                    if len(colunas) >= 3:
                        cnpj = colunas[0].get_text(strip=True)
                        razao_social = colunas[1].get_text(strip=True)
                        alteracao = colunas[2].get_text(strip=True)

                        if "CNPJ" not in cnpj and "Razão Social" not in razao_social and "Alteração" not in alteracao:
                            dados.append([cnpj, razao_social, alteracao, data_email])

    if dados:
        # Criar um DataFrame do pandas com a nova coluna "Data da Operação"
        df = pd.DataFrame(dados, columns=["CNPJ", "Razão Social", "Alteração", "Data da Operação"])

        # Garantir que os valores sejam strings
        df["CNPJ"] = df["CNPJ"].astype(str)
        df["Razão Social"] = df["Razão Social"].astype(str)
        df["Alteração"] = df["Alteração"].astype(str)

        # Adicionar um espaço não-quebrante e um espaço convencional
        df["CNPJ"] = df["CNPJ"].apply(lambda x: "\xa0 " + x if isinstance(x, str) else x)
        df["Razão Social"] = df["Razão Social"].apply(lambda x: "\xa0 " + x if isinstance(x, str) else x)
        df["Alteração"] = df["Alteração"].apply(lambda x: "\xa0 " + x if isinstance(x, str) else x)

        # Ordenar os dados do mais antigo para o mais novo
        df = df.sort_values(by="Data da Operação", ascending=True)

        # Remover colunas desnecessárias e garantir que todas as colunas sejam mantidas
        df = df.loc[:, ["CNPJ", "Razão Social", "Alteração", "Data da Operação"]]

        print(df)

        # Carregar o arquivo Excel e acessar a segunda planilha diretamente
        wb = load_workbook(caminho_excel)
        ws = wb.worksheets[1]  # A segunda aba, garantindo que seja "E-Mail BD"

        # Verificar se o nome da planilha corresponde ao esperado
        if ws.title != "E-Mail BD":
            print("⚠ A segunda planilha não tem o nome esperado. Verifique manualmente.")
        else:
            # Encontrar a primeira linha vazia
            primeira_linha_vazia = ws.max_row + 1

            # Inserir os novos dados
            for r_idx, row in enumerate(df.values, start=primeira_linha_vazia):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # **Preencher automaticamente as células E, F, G e H COM REFERÊNCIA DINÂMICA CORRETA**
            for coluna in ["E", "F", "G", "H"]:  # Percorre cada coluna
                ultima_linha_preenchida = primeira_linha_vazia - 1  # Última linha antes de colar os novos dados
                formula_origem = ws[f"{coluna}{ultima_linha_preenchida}"].value  # Captura a fórmula original da célula acima

                # Verifica se a célula contém uma fórmula
                if formula_origem and isinstance(formula_origem, str) and formula_origem.startswith("="):  
                    for linha in range(primeira_linha_vazia, ws.max_row + 1):
                        nova_formula = re.sub(r'([A-Z])(\d+)', lambda match: f"{match.group(1)}{linha}" if match.group(2) == str(ultima_linha_preenchida) else match.group(0), formula_origem)
                        ws[f"{coluna}{linha}"].value = nova_formula  # Aplica a fórmula ajustada para a linha correta

                else:
                    for linha in range(primeira_linha_vazia, ws.max_row + 1):
                        ws[f"{coluna}{linha}"].value = f"={coluna}{linha-1}"  

            # Aplicar alinhamento à coluna D (Data da Operação)
            for cell in ws["D"]:
                cell.alignment = Alignment(horizontal="center", vertical="top")

        # Salvar as alterações no arquivo Excel

            # Definir o novo nome do arquivo
            novo_nome_arquivo = f"Gerencie Carteira_{data_email.replace('/', '_')}.xlsx"

            # Definir o caminho para salvar o novo arquivo
            caminho_novo_excel = os.path.join(r"C:\Users\comercial05\Documents\Gerencie Carteira\Diário", novo_nome_arquivo)

            # Salvar como (com novo nome)
            wb.save(caminho_novo_excel)

            print(f"✅ Arquivo salvo como: {caminho_novo_excel}")

            print(f"✅ Dados copiados para a planilha 'E-Mail BD', com autopreenchimento das colunas E, F, G e H baseado na última linha acima.")
    else:
        print("⚠ Nenhuma tabela válida encontrada nos arquivos HTML.")
else:
    print("⚠ Nenhum e-mail não lido encontrado com o assunto especificado.")