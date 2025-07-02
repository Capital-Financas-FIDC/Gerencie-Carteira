#Mantém todas as funcionalidades da versão anterior

#Notas da v2.6.1:
#  - Adição das colunas I e J na planilha, para contabilizar as exclusões e inclusões

import win32com.client
import os
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re
from rich.console import Console
from rich.table import Table
from rich.panel import Panel

# Inicializa o objeto Console da biblioteca rich para impressões coloridas e formatadas no terminal.
console = Console()

# Conectar ao Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# Acessar a caixa de entrada (pasta Inbox)
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items

# Definir o assunto do e-mail para pesquisa
assunto_procurado = "Gerencie Carteira - Consulte as Empresas Monitoradas"

# Ordenar e-mails por data, do mais antigo ao mais recente
messages.Sort("[ReceivedTime]", False)

# Filtrar apenas os e-mails não lidos com o assunto especificado
emails_nao_lidos = [msg for msg in messages if msg.Subject == assunto_procurado and msg.UnRead]

# Verifica se há e-mails não lidos com o assunto procurado
if emails_nao_lidos:
    # Definir o caminho para a pasta onde os arquivos HTML serão salvos
    pasta_destino = r"C:\Users\comercial05\Documents\Gerencie Carteira\HTML"
    dados = []

    # Extrair a data do e-mail mais recente (embora não seja utilizada posteriormente para nomear o arquivo HTML individualmente,
    # a lógica para a data do Excel é mais robusta e independente da data do e-mail)
    data_mais_recente = max(email.ReceivedTime for email in emails_nao_lidos)

    # Definir o caminho para a pasta das planilhas diárias do Excel
    pasta_diario = r"C:\Users\comercial05\Documents\Gerencie Carteira\Diário"
    
    # Lista todos os arquivos na pasta 'Diário'
    arquivos_diario = os.listdir(pasta_diario)
    
    # Compila um padrão de expressão regular para filtrar arquivos Excel com o nome desejado e extrair a data
    # O padrão foi ajustado para AAAA_MM_DD
    padrao_nome_excel = re.compile(r"Gerencie Carteira_(\d{4}_\d{2}_\d{2})")
    
    datas_encontradas_lista = []

    # Itera sobre os arquivos na pasta 'Diário' para encontrar a data da última planilha existente
    for arquivo_diario in arquivos_diario:
        match = padrao_nome_excel.search(arquivo_diario)
        if match:
            datas_encontradas_lista.append(match.group(1))
        else:
            # Informa se nenhum arquivo com o padrão de data foi encontrado e abre a pasta para verificação manual
            console.print("[bold yellow]Nenhum arquivo com data foi encontrado na pasta! Verifique manualmente[/bold yellow]")
            os.startfile(pasta_diario)
            break

    # Se datas de arquivos Excel foram encontradas, determina a data da planilha mais recente
    if datas_encontradas_lista:
        # A chave de ordenação foi ajustada para lidar com o formato AAAA_MM_DD
        ultima_data_diario = max(datas_encontradas_lista, key=lambda x: tuple(map(int, x.split('_'))))
        # Constrói o nome completo do arquivo Excel mais recente
        nome_arquivo_excel = f"Gerencie Carteira_{ultima_data_diario}.xlsm"
        # Constrói o caminho completo para o arquivo Excel mais recente
        caminho_excel = os.path.join(pasta_diario, nome_arquivo_excel)
    else:
        # Informa se nenhuma planilha foi encontrada e abre a pasta para verificação manual
        console.print("[bold yellow]Nenhuma planilha encontrada na pasta! Verifique manualmente[/bold yellow]")
        os.startfile(pasta_diario)

    # Processa cada e-mail não lido encontrado
    for email in emails_nao_lidos:
        # Extrai a data do e-mail
        data_email = email.ReceivedTime.strftime("%d/%m/%Y") # Formata para a coluna "Data da Operação" como DD/MM/AAAA
        data_email_nome = email.ReceivedTime.strftime("%Y_%m_%d") # Formata para o nome do arquivo HTML/Excel como AAAA_MM_DD
        caminho_arquivo = None
        
        # Itera sobre os anexos do e-mail para encontrar o arquivo HTML
        for anexo in email.Attachments:
            if anexo.FileName.endswith(".html"):
                # Define o nome do arquivo HTML a ser salvo com a data formatada como AAAA_MM_DD
                nome_arquivo = f"Gerencie_Carteira_{data_email_nome}.html"
                # Constrói o caminho completo para salvar o anexo
                caminho_arquivo = os.path.join(pasta_destino, nome_arquivo)
                # Salva o anexo
                anexo.SaveAsFile(caminho_arquivo)
                console.print(f"Anexo salvo como: [green]'{nome_arquivo}'[/green] na pasta HTML's do Gerencie Carteira")
                print("\n")
                break

        # Se o arquivo HTML foi salvo, abre-o e extrai os dados da tabela
        if caminho_arquivo:
            with open(caminho_arquivo, "r", encoding="utf-8") as arquivo:
                soup = BeautifulSoup(arquivo, "html.parser")

            tabelas = soup.find_all("table")
            # Verifica se há mais de uma tabela e seleciona a segunda (índice 1)
            if len(tabelas) > 1:
                tabela = tabelas[1]
                tbody = tabela.find("tbody")
                # Encontra todas as linhas da tabela, priorizando 'tbody' se existir
                linhas = tbody.find_all("tr") if tbody else tabela.find_all("tr")

                # Itera sobre as linhas da tabela para extrair os dados
                for linha in linhas:
                    colunas = linha.find_all("td")
                    # Verifica se há pelo menos 3 colunas (CNPJ, Razão Social, Alteração)
                    if len(colunas) >= 3:
                        cnpj = colunas[0].get_text(strip=True)
                        razao_social = colunas[1].get_text(strip=True)
                        alteracao = colunas[2].get_text(strip=True)

                        # Adiciona os dados à lista se não forem os cabeçalhos da tabela
                        if "CNPJ" not in cnpj and "Razão Social" not in razao_social and "Alteração" not in alteracao:
                            dados.append([cnpj, razao_social, alteracao, data_email])

    # Se houver dados extraídos, processa-os e insere-os no Excel
    if dados:
        # Cria um DataFrame do pandas com os dados e a nova coluna "Data da Operação"
        df = pd.DataFrame(dados, columns=["CNPJ", "Razão Social", "Alteração", "Data da Operação"])
        
        # Converte todas as colunas para o tipo string
        df = df.astype(str)
        
        # Adiciona um espaço não-quebrante no início de cada valor para formatação
        df["CNPJ"] = df["CNPJ"].apply(lambda x: "\xa0 " + x)
        df["Razão Social"] = df["Razão Social"].apply(lambda x: "\xa0 " + x)
        df["Alteração"] = df["Alteração"].apply(lambda x: "\xa0 " + x)
        
        # Ordena os dados pela "Data da Operação" do mais antigo para o mais novo
        df = df.sort_values(by="Data da Operação", ascending=True)

        # Configura a tabela para exibição no terminal usando a biblioteca 'rich'
        tabela_rich = Table(title="Empresas Monitoradas", show_lines=True)
        tabela_rich.add_column("CNPJ", no_wrap=True, justify="center")
        tabela_rich.add_column("Razão Social", no_wrap=True, justify="center")
        tabela_rich.add_column("Alteração", justify="center")
        tabela_rich.add_column("Data da Operação", justify="center")

        # Popula a tabela 'rich' com os dados do DataFrame, aplicando formatação condicional à coluna "Alteração"
        for _, row in df.iterrows():
            cnpj_rich = str(row["CNPJ"]).strip()
            razao_social_rich = str(row["Razão Social"]).strip()
            alteracao_rich = str(row["Alteração"]).strip()
            data_rich = str(row["Data da Operação"]).strip()

            # Aplica cores diferentes para as alterações de inclusão, exclusão ou outras
            if alteracao_rich == "INCLUSAO  ANOT.INADIMPLENCIA":
                alteracao_rich_formatada = f"[bold red]{alteracao_rich}[/bold red]"
            elif alteracao_rich == "EXCLUSAO  ANOT.INADIMPLENCIA":
                alteracao_rich_formatada = f"[bold #1cb900]{alteracao_rich}[/bold #1cb900]"
            else:
                alteracao_rich_formatada = f"[bold yellow]{alteracao_rich}[/bold yellow]"
            
            tabela_rich.add_row(cnpj_rich, razao_social_rich, alteracao_rich_formatada, data_rich)

        # Imprime a tabela formatada no terminal
        console.print(tabela_rich)

        # Carrega o arquivo Excel
        wb = load_workbook(caminho_excel)
        # Acessa a segunda planilha (índice 1), que deve ser "E-Mail BD"
        ws_email_bd = wb.worksheets[1]
        # Acessa a quarta planilha (índice 3), que deve ser "PROCV GERENTES BD"
        ws_procv_gerentes_bd = wb.worksheets[2]

        # Verifica se o nome da segunda planilha corresponde ao esperado
        if ws_email_bd.title != "E-Mail BD":
            console.print("[bold red]A segunda planilha não tem o nome esperado. Verifique manualmente.[/bold red]")
            os.startfile(caminho_excel)
        # Verifica se o nome da quarta planilha corresponde ao esperado
        elif ws_procv_gerentes_bd.title != "PROCV GERENTES BD":
            console.print("[bold red]A terceira planilha não tem o nome esperado. Verifique manualmente.[/bold red]")
            os.startfile(caminho_excel)
        else:
            # Encontra a primeira linha vazia na planilha "E-Mail BD" para começar a inserir os novos dados
            primeira_linha_vazia_email_bd = ws_email_bd.max_row + 1

            # Insere os novos dados do DataFrame na planilha "E-Mail BD"
            for r_idx, row in enumerate(df.values, start=primeira_linha_vazia_email_bd):
                for c_idx, value in enumerate(row, start=1):
                    ws_email_bd.cell(row=r_idx, column=c_idx, value=value)

            # Preenche automaticamente as colunas E, F, G, H, I e J com fórmulas dinâmicas
            for coluna in ["E", "F", "G", "H", "I", "J"]:
                # Pega a última linha preenchida antes da inserção dos novos dados
                ultima_linha = primeira_linha_vazia_email_bd - 1
                # Captura a fórmula original da célula acima (na última linha preenchida)
                formula_origem = ws_email_bd[f"{coluna}{ultima_linha}"].value
                
                # Se a célula contém uma fórmula, ajusta-a para a nova linha
                if formula_origem and isinstance(formula_origem, str) and formula_origem.startswith("="):
                    for linha in range(primeira_linha_vazia_email_bd, ws_email_bd.max_row + 1):
                        # Substitui os números das linhas na fórmula para que apontem para a linha correta
                        nova_formula = re.sub(r'([A-Z])(\d+)', lambda m: f"{m.group(1)}{linha}" if m.group(2) == str(ultima_linha) else m.group(0), formula_origem)
                        ws_email_bd[f"{coluna}{linha}"].value = nova_formula
                else:
                    # Se não for uma fórmula, copia o valor da célula acima (com uma referência relativa simples)
                    for linha in range(primeira_linha_vazia_email_bd, ws_email_bd.max_row + 1):
                        ws_email_bd[f"{coluna}{linha}"].value = f"={coluna}{linha-1}"

            # Aplicar alinhamento centralizado e superior à coluna D (Data da Operação)
            for cell in ws_email_bd["D"]:
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # Verificar se cada Razão Social adicionada na coluna B está presente na coluna A da planilha "PROCV GERENTES BD"
            valores_procv = [cell.value for cell in ws_procv_gerentes_bd['A']]
            erros = False # Flag para detectar erros
            for linha in range(primeira_linha_vazia_email_bd, ws_email_bd.max_row + 1):
                valor = ws_email_bd[f"B{linha}"].value
                if valor not in valores_procv:
                    console.print(f"[bold red]Erro encontrado na célula E{linha}. Verifique manualmente.[/bold red]")
                    erros = True # Define a flag como True se um erro for detectado

            # Define o nome do novo arquivo Excel a ser salvo com a data formatada como AAAA_MM_DD
            novo_nome_arquivo = f"Gerencie Carteira_{data_email_nome}.xlsm"
            # Constrói o caminho completo para o novo arquivo Excel
            caminho_novo_excel = os.path.join(pasta_diario, novo_nome_arquivo)

            # Se algum erro foi detectado, salva o arquivo, exibe uma mensagem de erro e abre os arquivos relevantes
            if erros:
                wb.save(caminho_novo_excel)
                console.print(Panel.fit(f"Arquivo salvo como: [green]{novo_nome_arquivo}[/green]", title="Cedente Novo", border_style="yellow"))
                os.startfile(caminho_novo_excel)
                os.startfile(r"C:\DIRECIONA\atualiza.exe")
            # Se nenhum erro foi detectado, tenta atualizar a tabela dinâmica no Excel
            else:
                try:
                    # Acessa a primeira aba (índice 0), que deve ser "Tabela Dinâmica"
                    ws_tabela_dinamica = wb.worksheets[0]
                    # Define a propriedade para que a tabela dinâmica seja atualizada ao abrir o arquivo
                    pivot = ws_tabela_dinamica._pivots[0]
                    pivot.cache.refreshOnLoad = True
                    # Salva o arquivo Excel
                    wb.save(caminho_novo_excel)
                    console.print(Panel.fit(f"Tabela dinâmica atualizada com sucesso!\nArquivo salvo como: [green]{novo_nome_arquivo}[/green]", title="Sucesso", border_style="green"))
                except Exception as e:
                    # Informa se houve um erro ao tentar atualizar a tabela dinâmica
                    console.print(f"[bold red]Erro ao tentar atualizar a tabela dinâmica:[/bold red] {e}")
    else:
        # Informa se nenhuma tabela válida foi encontrada nos arquivos HTML
        console.print("[bold yellow]Nenhuma tabela válida encontrada nos arquivos HTML.[/bold yellow]")
        os.startfile(caminho_arquivo)
else:
    # Informa se nenhum e-mail não lido com o assunto especificado foi encontrado
    console.print("[bold yellow]Nenhum e-mail não lido encontrado com o assunto especificado.[/bold yellow]")

# Aguarda a entrada do usuário antes de sair para que as mensagens no terminal possam ser lidas
input("\nPressione ENTER para sair")